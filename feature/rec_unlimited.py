import speechRecognitionGoogle
import argparse
import tempfile
import queue
import sys
import datetime
import time as pf_time
import beep
import openpyxl
from udp.udp_receive import UDPReceive
import os
from dotenv import load_dotenv

import sounddevice as sd
import soundfile as sf
import numpy  # Make sure NumPy is loaded before it is used in the callback
assert numpy  # avoid "imported but unused" message (W0611)

load_dotenv()

class Recording:
    def __init__(self, filename=None, device=None, samplerate=None, channels=1, subtype=None):
        self.filename = filename
        self.device = device
        self.samplerate = samplerate
        self.channels = channels
        self.subtype = subtype
        self.q = queue.Queue()
        self.start_time = 0
        self.end_time = 0
        self.VOLUME_THRESHOLD = 10
        self.IS_RECORDING = False
        # self.udp_receive = UDPReceive(os.environ['MATSUKI7_IP'], 12345)
        self.udp_receive = UDPReceive('127.0.0.1', 12345)

    def int_or_str(self, text):
        """Helper function for argument parsing."""
        try:
            return int(text)
        except ValueError:
            return text

    def callback(self, indata, frames, time, status):
        """This is called (from a separate thread) for each audio block."""
        if status:
            print(status, file=sys.stderr)
        self.q.put(indata.copy())
        volume = numpy.linalg.norm(indata) * 10

        # if volume > self.VOLUME_THRESHOLD:
        #     self.IS_RECORDING = True

    def recording_to_text(self, reaction_time_sheet_path):
        try:
            if self.samplerate is None:
                device_info = sd.query_devices(self.device, 'input')
                # soundfile expects an int, sounddevice provides a float:
                self.samplerate = int(device_info['default_samplerate'])
            if self.filename is None:
                self.filename = tempfile.mktemp(prefix=datetime.datetime.now().strftime('%Y%m%d_%H%M%S'),
                                                suffix='.wav', dir='../sound/{}'.format(datetime.datetime.now().strftime('%Y%m%d')))

            beep.high()
            # datetime型に変換
            self.start_time = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S.%f')
            self.start_time = datetime.datetime.strptime(self.start_time, '%Y/%m/%d %H:%M:%S.%f')

            # Make sure the file is opened before recording anything:
            with sf.SoundFile(self.filename, mode='x', samplerate=self.samplerate,
                              channels=self.channels, subtype=self.subtype) as file:
                with sd.InputStream(samplerate=self.samplerate, device=self.device,
                                    channels=self.channels, callback=self.callback):
                    print('#' * 80)
                    print('ハンドルのボタンで発話を終了')
                    print('#' * 80)
                    while True:
                        # soundfileに書き込んでいる、writeはsoundfileのメソッド
                        # file.write(self.q.get())
                        # ハンドルのボタンが押されたら終了
                        # if self.udp_receive.is_finish_speaking(file,self.q):
                        #     self.end = pf_time.perf_counter()
                        #     print(self.end)
                        #     raise KeyboardInterrupt
                        if self.end_time == 0:
                            end_time_str = self.udp_receive.test()

                        if end_time_str is not None and self.end_time ==0:
                            # end_time_strをdatetime型に変換
                            self.end_time = datetime.datetime.strptime(end_time_str, '%Y/%m/%d %H:%M:%S.%f')

                            # 秒数にフォーマットしてself.end_timeとself.start_timeの差を取得
                            dif_time=self.end_time-self.start_time
                            # 秒数に変換
                            dif_time=dif_time.total_seconds()
                            print(dif_time)

                        if self.end_time !=0:
                            # ここが何か重い説？
                            file.write(self.q.get())
                            print(self.udp_receive.test_finish())
                            if self.udp_receive.test_finish():
                                raise KeyboardInterrupt

                        # if self.udp_receive.is_finish_speaking(file,self.q) and self.end_time !=0:
                        #     raise KeyboardInterrupt
                        
        except KeyboardInterrupt:
            beep.low()

            # excelシートを読み込む
            wb = openpyxl.load_workbook(reaction_time_sheet_path)
            sheet = wb.active
            # 最終行を取得
            last_row = sheet.max_row

            # 回数カラムに書き込む
            sheet.cell(row=last_row + 1, column=1, value=last_row)
            # reaction_timeカラムに書き込む
            sheet.cell(row=last_row + 1, column=2, value=self.end_time-self.start_time)
            print(self.end_time-self.start_time)

            # 保存
            wb.save(reaction_time_sheet_path)
            
            print('\nRecording finished: ' + repr(self.filename))
            text = speechRecognitionGoogle.speech_recognition(self.filename)
            return text
        except Exception as e:
            print(e)
