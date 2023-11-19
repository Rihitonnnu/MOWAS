import openpyxl
import os
import datetime

if os.path.exists('../data/token.xlsx'):
    write_wb = openpyxl.load_workbook('../data/token.xlsx')
if not os.path.exists('../data/token.xlsx'):
    write_wb = openpyxl.Workbook()
    write_wb.save('../data/token.xlsx')


class TokenRecord:
    def __init__(self):
        # シートを会話したときの日時で作成
        self.title = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        write_wb.create_sheet(title=self.title)

        # シートを変数に名前で格納
        self.write_ws = write_wb[self.title]

        c1 = self.write_ws["B1"]
        c2 = self.write_ws.cell(1, 3)
        c3 = self.write_ws.cell(1, 4)
        c1.value = "プロンプト"
        c2.value = "会話"
        c3.value = "合計"

        write_wb.save("../data/token.xlsx")

    # トークンを記録する
    def token_record(self, cb, cnt):
        data = [cnt, cb.prompt_tokens, cb.completion_tokens, cb.total_tokens]
        # シートを変数に名前で格納
        for i in range(1, 5):
            self.write_ws.cell(row=cnt+1, column=i, value=data[i-1])

        write_wb.save("../data/token.xlsx")
