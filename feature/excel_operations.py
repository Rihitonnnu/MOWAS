import openpyxl
from datetime import datetime

class ExcelOperations:
    def __init__(self, reaction_time_sheet_path):
        self.reaction_time_sheet_path = reaction_time_sheet_path

    # 反応時間を記録する
    def rac_time_excel(self, start_time, end_time,conv_start_time):
        # excelシートを読み込む
        wb = openpyxl.load_workbook(self.reaction_time_sheet_path)
        sheet = wb.active
        # 最終行を取得
        last_row = sheet.max_row
        # 回数カラムに書き込む
        sheet.cell(row=last_row + 1, column=1, value=last_row)
        # reaction_timeカラムに書き込む
        sheet.cell(row=last_row + 1, column=2, value=(end_time-start_time).total_seconds())
        # measurement_timeにself.conv_start_timeとself.end_timeの差分を書き込む
        sheet.cell(row=last_row + 1, column=3, value=(end_time-conv_start_time))

        print(end_time-start_time)
        # 保存
        wb.save(self.reaction_time_sheet_path)
    
    # 案内受容時間を記録する
    def guide_acc_time_excel(self, conv_start_time):
        # excelシートを読み込む
        wb = openpyxl.load_workbook(self.reaction_time_sheet_path)
        sheet = wb.active

        guide_acc_time=datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S.%f')
        guide_acc_time=datetime.datetime.strptime(guide_acc_time,'%Y/%m/%d %H:%M:%S.%f')
        # 回数カラムに書き込む
        sheet.cell(row=2, column=4, value=guide_acc_time-conv_start_time)
        # 保存
        wb.save(self.reaction_time_sheet_path)
